#!/usr/bin/env python3
"""Convert the public Excel registry into a compact, quality-checked browser data file."""

from __future__ import annotations

import base64
import gzip
import json
import math
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


REQUIRED_HEADERS = {
    "Nr rendor",
    "Emri tregtar",
    "Substanca aktive",
    "ATC Code",
    "Klasa / Çka është",
    "Përdorimi (fjalë kyçe)",
}

QUALITY_VERSION = "2026-07-22.1"
REGISTRY_CORRECTIONS: tuple[dict[str, Any], ...] = (
    {
        "id": "REG-2026-001",
        "match": {"ProtocolNo": "PD1339/051225", "PDID": "42"},
        "set": {"Substanca aktive": "Metamizole sodium"},
        "reason": (
            "Burimi e kishte substancën gabimisht si Metronidazole micronised; "
            "ANALGIN 1 g/2 ml me ATC N02BB02 përputhet me metamizole sodium."
        ),
        "sourceUrl": "https://lekovi.zdravstvo.gov.mk/drugsregister/detailview/51155",
        "verifiedAt": "2026-07-22",
    },
)


def normalize(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and math.isnan(value):
        return ""
    return value


def text(value: Any) -> str:
    return str(value if value is not None else "").strip()


def clean_headers(row: Iterable[Any]) -> list[str]:
    return [text(value) for value in row]


def find_header_row(sheet) -> tuple[list[str], int]:
    """Find the real header row within the first 15 rows."""
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=1, max_row=15, values_only=True),
        start=1,
    ):
        headers = clean_headers(row)
        if REQUIRED_HEADERS.issubset(set(headers)):
            return headers, row_number

    raise SystemExit(
        "Could not find the registry header row. Required columns: "
        + ", ".join(sorted(REQUIRED_HEADERS))
    )


def correction_matches(record: dict[str, Any], correction: dict[str, Any]) -> bool:
    return all(text(record.get(field)) == text(expected) for field, expected in correction["match"].items())


def apply_quality_corrections(records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    applied: list[dict[str, Any]] = []

    for record in records:
        for correction in REGISTRY_CORRECTIONS:
            if not correction_matches(record, correction):
                continue
            original = {field: record.get(field, "") for field in correction["set"]}
            record.update(correction["set"])
            record["__qualityStatus"] = "corrected"
            record["__qualityMessage"] = f"{correction['id']}: {correction['reason']}"
            record["__qualitySourceUrl"] = correction["sourceUrl"]
            record["__qualityVersion"] = QUALITY_VERSION
            applied.append({
                "id": correction["id"],
                "ProtocolNo": record.get("ProtocolNo", ""),
                "PDID": record.get("PDID", ""),
                "original": original,
                "corrected": correction["set"],
            })

    return records, applied


def assert_critical_registry_rules(records: list[dict[str, Any]]) -> None:
    matches = [
        record for record in records
        if text(record.get("ProtocolNo")) == "PD1339/051225"
        and text(record.get("PDID")) == "42"
    ]
    if len(matches) != 1:
        raise SystemExit(
            "Critical quality rule failed: expected exactly one ANALGIN PD1339/051225 / PDID 42 row."
        )

    analgin = matches[0]
    if text(analgin.get("Substanca aktive")).lower() != "metamizole sodium":
        raise SystemExit(
            "Critical quality rule failed: ANALGIN PD1339/051225 must be Metamizole sodium."
        )
    if text(analgin.get("ATC Code")).upper().replace(" ", "") != "N02BB02":
        raise SystemExit(
            "Critical quality rule failed: ANALGIN PD1339/051225 must have ATC N02BB02."
        )

    for record in records:
        active = text(record.get("Substanca aktive")).lower()
        atc = text(record.get("ATC Code")).upper().replace(" ", "")
        if "metronidazole" in active and atc == "N02BB02":
            raise SystemExit(
                "Critical quality rule failed: metronidazole cannot be published under ATC N02BB02."
            )


def read_records(source: Path) -> tuple[list[dict[str, Any]], str, int]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook.active
    headers, header_row_number = find_header_row(sheet)

    records: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=header_row_number + 1, values_only=True):
        if not any(value is not None and text(value) for value in row):
            continue

        record = {
            header: normalize(value)
            for header, value in zip(headers, row)
            if header
        }
        records.append(record)

    if not records:
        raise SystemExit("No medicine rows were found below the header.")

    return records, sheet.title, header_row_number


def main() -> None:
    if len(sys.argv) != 3:
        raise SystemExit("Usage: sync_sheet.py SOURCE.xlsx OUTPUT.js")

    source = Path(sys.argv[1])
    output = Path(sys.argv[2])

    records, sheet_title, header_row_number = read_records(source)
    records, applied = apply_quality_corrections(records)
    assert_critical_registry_rules(records)

    raw_json = json.dumps(
        records,
        ensure_ascii=False,
        separators=(",", ":"),
    ).encode("utf-8")
    packed = base64.b64encode(gzip.compress(raw_json, compresslevel=9)).decode("ascii")
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        "window.DRUG_DATA_PARTS = [\"" + packed + "\"];\n"
        + "window.REGISTRY_STATIC_META = "
        + json.dumps(
            {
                "count": len(records),
                "generatedAt": generated_at,
                "qualityVersion": QUALITY_VERSION,
                "correctionsApplied": len(applied),
            },
            ensure_ascii=False,
            separators=(",", ":"),
        )
        + ";\n",
        encoding="utf-8",
    )

    print(f"Synced {len(records)} medicines from sheet '{sheet_title}'.")
    print(f"Header row: {header_row_number}.")
    print(f"Quality version: {QUALITY_VERSION}; corrections applied: {len(applied)}.")
    print(f"Generated {output} ({output.stat().st_size} bytes).")


if __name__ == "__main__":
    main()
